import logging
import re
import zipfile

import openpyxl

from converter_app.readers.helper.base import Reader
from converter_app.readers.helper.reader import Readers

logger = logging.getLogger(__name__)


class SuprabankReader(Reader):
    """
    Reads JASCO FP-8300 fluorescence titration exports converted to xlsx by the Suprabank platform.

    Layout of the single worksheet:
      row 1: 'Wavelength' | 'Intensity' (raw block) | ... | 'Intensity' (corrected block)
      row 2: units 'nm' | 'a.u.' | ...
      row 3: per-column sample labels, e.g. '11 uL Phenylethylamine'
      row 4: dilution-correction formulas (documentation only) for the corrected block
      rows 5..N: numeric data, column A = wavelength, following columns = intensity per tube
      rows N+2..: '##### Extended Information' metadata block (key in column A, value in column B)

    It emits three families of tables:
      - one metadata table built from the 'Extended Information' block
      - spectra pages: one table per tube (x = wavelength, y = intensity), per block
      - titration pages: one table per wavelength (x = volume, y = intensity), per block
    """
    identifier = 'suprabank_reader'
    # keep below ExcelReader (priority 15) so this specific format wins over the generic Excel reader
    priority = 14

    # marker that separates the numeric data from the trailing metadata block
    _metadata_marker = '##### Extended Information'
    _volume_pattern = re.compile(r'(-?\d+(?:[.,]\d+)?)')

    def __init__(self, file, *tar_content):
        super().__init__(file, *tar_content)
        self.wb = None
        self.ws = None

    def check(self):
        """
        :return: True if the file is a Suprabank JASCO fluorescence xlsx export
        """
        if self.file.encoding != 'binary' or self.file.suffix != '.xlsx':
            return False

        try:
            self.wb = openpyxl.load_workbook(filename=self.file.fp)
        except (openpyxl.utils.exceptions.InvalidFileException, zipfile.BadZipFile):
            return False

        self.ws = self.wb.active

        # verify the characteristic header signature
        header = [str(c.value).strip() if c.value is not None else '' for c in self.ws[1]]
        units = [str(c.value).strip() if c.value is not None else '' for c in self.ws[2]]
        if not header or header[0] != 'Wavelength' or 'Intensity' not in header:
            return False
        if not units or units[0] != 'nm':
            return False

        # verify the trailing 'Extended Information' metadata block is present
        has_marker = any(
            row[0] is not None and str(row[0]).strip() == self._metadata_marker
            for row in self.ws.iter_rows(min_col=1, max_col=1, values_only=True)
        )
        result = has_marker
        logger.debug('result=%s', result)
        return result

    def prepare_tables(self):
        rows = list(self.ws.iter_rows(values_only=True))

        header = [self._clean(v) for v in rows[0]]
        units = [self._clean(v) for v in rows[1]]
        labels = [self._clean(v) for v in rows[2]]
        formulas = [self._clean(v) for v in rows[3]]

        # split the intensity columns into blocks (raw / corrected), each starting at an 'Intensity' header cell
        blocks = self._detect_blocks(header, labels, formulas)

        # collect the numeric data rows (column A holds the wavelength)
        data_rows = []
        meta_start = None
        for idx in range(4, len(rows)):
            first = rows[idx][0]
            if str(first).strip() == self._metadata_marker:
                meta_start = idx
                break
            if isinstance(first, (int, float)):
                data_rows.append(rows[idx])

        wavelengths = [row[0] for row in data_rows]
        x_unit = units[0] if units else 'nm'

        tables = []

        # 1) metadata table from the 'Extended Information' block
        meta_table = self.append_table(tables)
        if meta_start is not None:
            for idx in range(meta_start + 1, len(rows)):
                key = self._clean(rows[idx][0])
                value = self._clean(rows[idx][1]) if len(rows[idx]) > 1 else None
                if key and value is not None:
                    meta_table.add_metadata(str(key), str(value))
        meta_table['metadata']['internal_reader_name'] = 'suprabank'

        # 2) spectra pages: one table per tube -> x = wavelength, y = intensity
        for block_name, cols in blocks:
            volumes = [self._parse_volume(labels[c]) for c in cols]
            y_unit = units[cols[0]] if len(units) > cols[0] and units[cols[0]] else 'a.u.'
            for col, volume in zip(cols, volumes):
                table = self.append_table(tables)
                table['metadata']['internal_reader_name'] = 'suprabank'
                table['metadata']['internal_reader_type'] = 'fluorescence spectrum'
                table['metadata']['block'] = block_name
                table['metadata']['Volume'] = f'{volume} uL'
                table['metadata']['Sample'] = str(labels[col])
                table['metadata']['AllVolumes'] = str(volumes)
                table['columns'] = [
                    {'key': '0', 'name': f'Wavelength [{x_unit}]'},
                    {'key': '1', 'name': f'Intensity [{y_unit}]'},
                ]
                table['rows'] = [[row[0], row[col]] for row in data_rows]

        # 3) titration pages: one table per wavelength -> x = volume, y = intensity
        for block_name, cols in blocks:
            volumes = [self._parse_volume(labels[c]) for c in cols]
            y_unit = units[cols[0]] if len(units) > cols[0] and units[cols[0]] else 'a.u.'
            for row, wavelength in zip(data_rows, wavelengths):
                table = self.append_table(tables)
                table['metadata']['internal_reader_name'] = 'suprabank'
                table['metadata']['internal_reader_type'] = 'titration'
                table['metadata']['block'] = block_name
                table['metadata']['Wavelength'] = f'{wavelength} {x_unit}'
                table['metadata']['AllWaves'] = str(wavelengths)
                table['columns'] = [
                    {'key': '0', 'name': 'Volume [uL]'},
                    {'key': '1', 'name': f'Intensity [{y_unit}]'},
                ]
                table['rows'] = [[volume, row[col]] for volume, col in zip(volumes, cols)]

        return tables

    def _detect_blocks(self, header, labels, formulas):
        """
        Groups the intensity columns into blocks. A block starts at every 'Intensity' header cell
        and reaches until the next one. A block is 'corrected' when its columns carry formula
        annotations in row 4, otherwise 'raw'.
        """
        starts = [idx for idx, value in enumerate(header) if value == 'Intensity']
        blocks = []
        for i, start in enumerate(starts):
            end = starts[i + 1] if i + 1 < len(starts) else len(header)
            cols = [c for c in range(start, end) if c < len(labels) and labels[c] is not None]
            if not cols:
                continue
            is_corrected = any(c < len(formulas) and formulas[c] is not None for c in cols)
            blocks.append(('corrected' if is_corrected else 'raw', cols))
        return blocks

    def _parse_volume(self, label):
        """Extracts the leading numeric volume from a sample label like '11 uL Phenylethylamine'."""
        if label is None:
            return ''
        match = self._volume_pattern.match(str(label).strip())
        return match.group(1) if match else str(label)

    @staticmethod
    def _clean(value):
        """Strips strings, leaves other values untouched, maps blanks to None."""
        if isinstance(value, str):
            value = value.strip()
            return value if value else None
        return value


Readers.instance().register(SuprabankReader)
